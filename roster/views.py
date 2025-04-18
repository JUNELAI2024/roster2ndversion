from django.shortcuts import render, get_object_or_404, redirect
from rest_framework import viewsets
from .models import Staff, Roster, RosterConfig, BakeryProduct, BakeryProductRestock, DailyRevenue, UserAccount
from django.db.models import Sum, Count, Q
from .serializers import StaffSerializer, RosterSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.utils import timezone
from django.contrib import messages
from datetime import  timedelta, datetime 
from django import forms
import logging
logger = logging.getLogger(__name__)
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models.functions import TruncDate
from decimal import Decimal
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.template.loader import get_template
from xhtml2pdf import pisa
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


@login_required
def modify_product_info(request):
    product = None
    if request.method == "POST":
        item_id = request.POST.get('item_id')
        item_name = request.POST.get('item_name')

        # Search for the product based on item_id or item_name
        if item_id:
            product = get_object_or_404(BakeryProduct, item_id=item_id)
        elif item_name:
            product = BakeryProduct.objects.filter(item_name__icontains=item_name).first()

        # If a product is found and the modify button is pressed
        if product and 'modify' in request.POST:
            product.item_name = request.POST.get('item_name', product.item_name)
            product.item_name_CHI = request.POST.get('item_name_CHI', product.item_name_CHI)
            product.category = request.POST.get('category', product.category)
            product.remarks = request.POST.get('remarks', product.remarks)
            product.save()  # Save updated product information
            return redirect('modify_product_info')  # Redirect after saving

    return render(request, 'roster/modify_product_info.html', {'product': product})

@login_required
def manage_bakery_products(request):
    if request.method == 'POST':
        # Handle the form submission
        item_id = request.POST.get('item_id')
        category = request.POST.get('category')
        item_name = request.POST.get('item_name')
        item_name_CHI = request.POST.get('item_name_CHI', '')
        onsell = request.POST.get('onsell') == 'True'
        start_date = request.POST.get('start_date')
        shelved_date = request.POST.get('shelved_date') or None
        remarks = request.POST.get('remarks', '')
        image_file = request.FILES.get('image_url')  # Get the uploaded image file

         # Save the image to the media directory
        if image_file:
            # Define the path to save the image directly in MEDIA_ROOT
            media_path = settings.MEDIA_ROOT  # Use MEDIA_ROOT directly
            
            # Define the image name and full path
            image_name = image_file.name  # Get the image name
            image_full_path = os.path.join(media_path, image_name)  # Full path to save the image
            
            # Save the image file
            with open(image_full_path, 'wb+') as destination:
                for chunk in image_file.chunks():
                    destination.write(chunk)

            # Generate the URL for the image
            image_url = f'/roster/media/{image_name}'  # Custom URL format

        else:
            image_url = ''  # Handle case where no image is uploaded

        # Create a new BakeryProduct instance
        new_product = BakeryProduct(
            item_id=item_id,
            category=category,
            item_name=item_name,
            item_name_CHI=item_name_CHI,
            onsell=onsell,
            start_date=start_date,
            shelved_date=shelved_date,
            remarks=remarks,
            image_url=image_url  # Store the generated URL
        )
        new_product.save()  # Save to the database

        return redirect('manage_bakery_products')  # Redirect to the same page after submission

    # Handle GET request to display the form
    return render(request, 'roster/manage_bakery_products.html')

@login_required
def bakery_product_view(request):
    products = BakeryProduct.objects.filter(onsell=True)  # Get only products that are on sale
    categories = products.values_list('category', flat=True).distinct()  # Get unique categories

    context = {
        'products': products,
        'categories': categories,
    }
    
    return render(request, 'roster/bakery_product.html', context)

@login_required
@csrf_exempt
def restock_product(request):
     # Fetch all products that are onsell
    products = BakeryProduct.objects.filter(onsell=True)
    # Get unique categories
    categories = products.values_list('category', flat=True).distinct()

    

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            for product in data:
               item_id = product.get('item_id')
            product_name = product.get('product_name')
            restock_quantity = product.get('restock_quantity')
            delivery_date = product.get('delivery_date')
            order_by = product.get('order_by')

            # Create a new instance and save it to the database
            restock_entry = BakeryProductRestock(
                item_id=item_id,
                product_name=product_name,
                restock_quantity=restock_quantity,
                delivery_date=delivery_date,
                order_by=order_by
            )
            restock_entry.save()  # update_date will be set automatically here

            return JsonResponse({'status': 'success', 'message': 'Product restocked successfully.'})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
        
 # Handle GET requests by rendering the template
    context = {
        'categories': categories,
         'products': products  # Pass products to the template, including image URLs
    }
    # Handle GET requests by rendering the template
    return render(request, 'roster/bakery_product.html')  # Ensure this points to your actual template path


def home(request):
    return render(request, 'roster/home.html')  # Make sure the path matches your template location

# Original views for rendering templates
@login_required
def staff_list(request):
    active_staff = Staff.objects.filter(is_active=True)  # Get only active staff
    return render(request, 'roster/staff_list.html', {'staff_list': active_staff})

@login_required
def roster_create(request):
    active_staff = Staff.objects.filter(is_active=True)  # Get active staff
    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']  # Define days of the week
    week_start_date = None  # Initialize week_start_date
    time_slots = RosterConfig.objects.values_list('time_slot', flat=True).order_by('time_slot')  # Fetch time slots
    formatted_time_slots = [slot.strftime('%H:%M') for slot in time_slots]  # Format to HH:MM:SS
    duty_roles = RosterConfig.objects.values_list('duty_role', flat=True).distinct().order_by('duty_role')
 
    submitted_data = []  # Initialize here to ensure it's always defined

    if request.method == 'POST':
        # Get the week starting date from the form input
        week_start_date_str = request.POST.get('week_start_date')
        if week_start_date_str:  # If the user provided a date
            week_start_date = datetime.strptime(week_start_date_str, '%Y-%m-%d').date()  # Convert to date
        else:
            week_start_date = timezone.now().date()  # Fallback to today's date if not provided
       

        for staff in active_staff:
            staff_row = [staff.name]  # Store the staff member's name
            for day in days_of_week:
                shift_start = request.POST.get(f"shift_start_{staff.id}_{day}")
                shift_end = request.POST.get(f"shift_end_{staff.id}_{day}")
                duty_role = request.POST.get(f"duty_role_{staff.id}_{day}")  # Get duty role name directly

                if shift_start and shift_end:
                    # Calculate the work date based on the week start date and day
                    work_date = week_start_date + timedelta(days=days_of_week.index(day))

                    # Check if a roster entry already exists
                    if Roster.objects.filter(staff_name=staff.name, day=day, work_date=work_date).exists():
                        messages.error(request, f"Shift for {staff.name} on {day} already exists.")
                        continue

                    # Convert to TimeField directly
                    shift_start_time = timezone.datetime.strptime(shift_start, "%H:%M").time()
                    shift_end_time = timezone.datetime.strptime(shift_end, "%H:%M").time()
                    no_of_work_hr = round((timezone.datetime.combine(work_date, shift_end_time) - 
                                            timezone.datetime.combine(work_date, shift_start_time)).seconds / 3600.0, 1)


                    # Create roster entry
                    Roster.objects.create(
                        staff_name=staff.name,  # Store the staff member's name
                        day=day,
                        shift_start=shift_start_time,
                        shift_end=shift_end_time,
                        duty_role=duty_role,  # Store the duty role name directly
                        week_start_date=week_start_date,
                        work_date=work_date,
                        no_of_work_hr=no_of_work_hr,  # Store the calculated working hours
                       
                    )

 # Collect data for Excel generation
            # Format for Excel: "Start - End (Role)"
                    combined_entry = f"{shift_start} - {shift_end} ({duty_role})"
                    staff_row.append(combined_entry)
                else:
                    staff_row.append('')  # Empty cell if no shift

            submitted_data.append(staff_row)


        messages.success(request, "Roster created successfully!")

          # Generate the Excel file using submitted data
      # Call the function to generate the roster and send the email
        generate_roster_excel_file(request, week_start_date, submitted_data)

        return redirect('roster_list')

    return render(request, 'roster/roster_create.html', {
        'staff_list': active_staff,
        'days': days_of_week,
        'week_start_date': week_start_date,
        'time_slots': formatted_time_slots ,  # Pass time slots directly
        'duty_roles': duty_roles,  # Pass duty roles to the template
    })



def generate_roster_excel_file(request, week_start_date, submitted_data):
    # Create a Workbook and add data
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Add a main title header
    main_title = f"Roster - {week_start_date.strftime('%Y-%m-%d')}"
    ws.append([main_title])  # Add the main title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(submitted_data[0]) + 1)  # Merge cells for title
    
    # Style the main title
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Prepare the header
    header = ['Staff Member']
    days_of_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for day in days_of_week:
        header.append(day)  # Single column for combined data

    ws.append(header)

    # Apply styles to the header
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
    for cell in ws[2]:  # Second row is the header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill  # Apply fill color

    # Add the submitted data to the worksheet
    for staff_row in submitted_data:
        ws.append(staff_row)

    # Optional: Apply styling to data rows (e.g., alternate row colors)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(header)):
        if (row[0].row % 2) == 0:  # Even rows
            fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Light grey fill
        else:  # Odd rows
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill

        for cell in row:
            cell.fill = fill  # Apply fill color

    # Set column widths
    column_widths = [20] + [25] * 7  # Adjust widths as needed
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Generate a timestamp for the filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = f'media/roster_{week_start_date.strftime("%Y%m%d")}_{timestamp}.xlsx'
    wb.save(file_path)


@login_required
def roster_list(request):
    logger.info(f"Session ID: {request.session.session_key}")
    logger.info(f"User is authenticated: {request.user.is_authenticated}")
    active_staff = Staff.objects.filter(is_active=True)  # Fetch only active staff
    roster_list = Roster.objects.all()  # Initial query

    staff_name = request.GET.get('staff_name') 
    month = request.GET.get('month')

    if staff_name:  # Adjusted to use staff_name
        roster_list = roster_list.filter(staff_name__iexact=staff_name)  # Filter by staff name (case-insensitive)

    if month:
        roster_list = roster_list.filter(work_date__month=month)  # Assuming work_date is a DateField

    context = {
        'active_staff': active_staff,
        'roster_list': roster_list,
    }
    return render(request, 'roster/roster_list.html', context)


# New view for displaying shift statistics
@login_required
def statistics_view(request):
    # Calculate total hours worked by summing the no_of_work_hr field
    total_hours_worked = Roster.objects.aggregate(total_hours=Sum('no_of_work_hr'))['total_hours'] or 0.0
     # Calculate total working hours by day
    daily_hours = (
        Roster.objects
        .annotate(date=TruncDate('work_date'))  # Group by day
        .values('date')  # Get date values
        .annotate(total_hours=Sum('no_of_work_hr'))  # Sum hours for each day
        .order_by('date')  # Order by date
    )

    # Prepare data for chart
    dates = [entry['date'].strftime('%Y-%m-%d') for entry in daily_hours]  # Format dates
    total_daily_hours = [entry['total_hours'] for entry in daily_hours]  # Get total hours for each day


     # Calculate total working hours by staff
    staff_hours = (
        Roster.objects
        .values('staff_name')  # Group by staff name
        .annotate(total_hours=Sum('no_of_work_hr'))  # Sum hours for each staff
        .order_by('-total_hours')  # Order by total hours descending
    )

    # Convert staff_hours to a list of dictionaries for easier access in the template
    staff_hours_list = list(staff_hours)
 # Get the top 5 staff by total working hours
    top_staff = staff_hours[:5]  # Get the top 5 staff members


    return render(request, 'roster/statistics.html', {
        'total_hours_worked': total_hours_worked,
        'total_daily_hours': total_daily_hours,
        'dates': json.dumps(dates),  # Convert dates to JSON
        'staff_hours': json.dumps(staff_hours_list),  # Convert staff hours to JSON
        'top_staff': json.dumps(list(top_staff)),  # Convert top staff to JSON
    })

# New API view to get shift counts similar to statistics_view
@api_view(['GET'])
def api_shift_counts(request):
    shift_counts = (
        Roster.objects
        .values('staff__name')
        .annotate(
            am_count=Count('id', filter=Q(shift_start='09:00:00')),
            pm_count=Count('id', filter=Q(shift_start='14:00:00')),
            full_count=Count('id', filter=Q(shift_start='09:00:00', shift_end='19:00:00'))
        )
    )

    return Response(shift_counts)

# New API views using Django REST Framework
class StaffViewSet(viewsets.ModelViewSet):
    queryset = Staff.objects.all()
    serializer_class = StaffSerializer

class RosterViewSet(viewsets.ModelViewSet):
    queryset = Roster.objects.all()
    serializer_class = RosterSerializer



@login_required
def submit_revenue(request):
    if request.method == 'POST':
        try:
            user_input = request.POST.get('user')
            business_date = request.POST.get('businessDate')
            business_time = request.POST.get('businessTime')
            amex = Decimal(request.POST.get('amex', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            debit_card = Decimal(request.POST.get('debit', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            visa = Decimal(request.POST.get('visa', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            mastercard = Decimal(request.POST.get('mastercard', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            cash = Decimal(request.POST.get('cash', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            unionpay = Decimal(request.POST.get('unionpay', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            wonderful_card = Decimal(request.POST.get('wonderfulCard', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            gift_card = Decimal(request.POST.get('giftCard', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            pst = Decimal(request.POST.get('pst', 0) or 0).quantize(Decimal('0.01'))  # Standard rounding
            redeem_points = int(request.POST.get('redeemPoints', 0))

            # Calculate total
            total = (amex + debit_card + visa + mastercard + cash + unionpay + wonderful_card).quantize(Decimal('0.01'))  # Standard rounding
            # Calculate total sum
            total_sum = (float(amex) + float(debit_card) + float(visa) + 
                     float(mastercard) + float(cash) + 
                     float(unionpay) + float(wonderful_card) + 
                     float(gift_card) + float(pst) + 
                     float(redeem_points))
            
            # Create and save the revenue record
            DailyRevenue.objects.create(
                user_input=user_input,  # Make sure this is set correctly
                business_date=business_date,
                business_time=business_time,
                amex=amex,
                debit_card=debit_card,
                visa=visa,
                mastercard=mastercard,
                cash=cash,
                unionpay=unionpay,
                wonderful_card=wonderful_card,
                gift_card=gift_card,
                pst=pst,
                redeem_points=redeem_points,
                total_sum=total_sum,  # Store the total sum
                total=total
            )

            return JsonResponse({'status': 'success', 'message': 'Revenue submitted successfully!'})
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return render(request, 'roster/submit_revenue.html')  # Adjust to your template name

@login_required
def revenue_dashboard(request):
    # Sample data; replace with your actual data retrieval logic
    total_revenue = DailyRevenue.objects.aggregate(TotalSum=Sum('total_sum'))['TotalSum'] or 0.00
    # Count the number of unique business dates
    number_of_days = DailyRevenue.objects.values('business_date').distinct().count()
     # Calculate average revenue per day
    average_revenue_per_day = total_revenue / number_of_days if number_of_days > 0 else 0.00
    # Calculate total revenue for each payment method
    payment_methods = [
        'amex', 'debit_card', 'visa', 'mastercard',
        'cash', 'unionpay', 'wonderful_card'
    ]
    payment_data = {}
    for method in payment_methods:
        total = DailyRevenue.objects.aggregate(total_sum=Sum(method))['total_sum'] or 0.0
        payment_data[method] = float(total)  # Ensure values are floats

 # Find the top payment method
    # top_payment_method = max(payment_data, key=payment_data.get)
    # top_payment_value = payment_data[top_payment_method]

    # Get the entry with the highest total_sum
    top_revenue_entry = DailyRevenue.objects.order_by('-total_sum').first()
    top_revenue_value = top_revenue_entry.total_sum if top_revenue_entry else 0
    top_revenue_day_date = top_revenue_entry.business_date if top_revenue_entry else 'N/A'  # Use business_date
   
     # Prepare data for the chart
    payment_method_labels = list(payment_data.keys())
    payment_method_values = list(payment_data.values())
    

    # Unique business dates and their total sums
    revenue_over_time = DailyRevenue.objects.values('business_date').annotate(total_sum=Sum('total_sum')).order_by('business_date')

     # Prepare labels and data for the chart
    labels = [entry['business_date'].strftime('%Y-%m-%d') for entry in revenue_over_time]
    revenue_data = [float(entry['total_sum']) for entry in revenue_over_time]  # Ensure values are floats


    context = {
        'total_revenue': total_revenue,
        'average_revenue_per_day': average_revenue_per_day,
        'revenue_data': revenue_data,
        'labels': labels,
        'payment_method_labels': payment_method_labels,
        'payment_method_values': payment_method_values,
         #'top_payment_method': top_payment_method,
        #'top_payment_value': top_payment_value,
        'top_revenue_value': top_revenue_value,
        'top_revenue_day_date': top_revenue_day_date,
    }

    return render(request, 'roster/revenue_dashboard.html', context)


def login_view(request):
    # Check if the user is already authenticated
    if request.user.is_authenticated:
        return redirect('revenue_dashboard')  # Redirect to the dashboard if already logged in

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('revenue_dashboard')  # Redirect to the dashboard after login
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'roster/home.html')

def logout_view(request):
    logout(request)
    return redirect('home')  # Redirect to home after logging out


@login_required
def manage_staff(request, staff_id=None):
    if staff_id:
        staff = get_object_or_404(Staff, id=staff_id)
    else:
        staff = Staff()

    if request.method == 'POST':
        name = request.POST.get('name')
        position = request.POST.get('position')
        is_active = request.POST.get('is_active') == 'on'  # Checkbox returns 'on' when checked

        if staff_id:
            staff.name = name
            staff.position = position
            staff.is_active = is_active
            staff.save()
        else:
            Staff.objects.create(name=name, position=position, is_active=is_active)

        return redirect('manage_staff')  # Redirect to staff management page or list

    return render(request, 'roster/manage_staff.html', {'staff': staff})

@login_required
def export_report(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        format_type = request.POST.get('format')

        if report_type == 'roster':
            data = Roster.objects.all()
        elif report_type == 'staff':
            data = Staff.objects.all()

            # Debugging
        print("Data fetched for report type:", report_type)
        print("Data:", list(data))  # Print the data to check its structure

        if format_type == 'excel':
            df = pd.DataFrame(list(data.values()))
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
            df.to_excel(response, index=False)
            return response

        elif format_type == 'pdf':
            template_path = 'roster/report_template.html'
            context = {
                'report_type': report_type,
                'roster_data': data if report_type == 'roster' else [],
                'staff_data': data if report_type == 'staff' else [],
            }
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
            template = get_template(template_path)
            html = template.render(context)
            pisa_status = pisa.CreatePDF(html, dest=response)

            if pisa_status.err:
                return HttpResponse('Error generating PDF', status=400)

            return response

    return render(request, 'roster/export_report.html')

@login_required
def product_list(request):
    # Retrieve products with non-empty image_url
    products = BakeryProduct.objects.exclude(image_url__isnull=True).exclude(image_url__exact='')
    
    return render(request, 'roster/product_list.html', {'products': products})